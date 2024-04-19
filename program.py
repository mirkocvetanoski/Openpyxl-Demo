from openpyxl import *
from openpyxl.styles import Font

# Filter names
filter_names = [
    "sheet name",
    "no.",
    "description",
    "unit",
    "boq qty",
    "unit price",
    "amount",
    "prev qty",
    "this qty",
    "total qty",
    "prev amount",
    "this amount",
    "cumulative",
    "%",
    "custom1",
    "custom2",
    "custom3",
    "custom4",
    "custom5",
    "custom6",
    "price form",
    "contract or vo",
    "item",
    "object",
]

def populate_filter_names(sheet):
    for name in filter_names:
        sheet.cell(1, filter_names.index(name) + 1).value = name
        sheet.cell(1, filter_names.index(name) + 1).font = Font(bold=True)

def populate_cells(dataSheet, counter, col_num, sheet, cell):
    dataSheet.cell(counter, col_num).value = f"={sheet.title}!{cell.coordinate}"

def save_wb(wb, filename):
    # Save a workbook
    wb.save(filename)

def main():
    filename = "Granit.xlsx"

    #Open the workbook
    wb = load_workbook(filename)

    if 'Data' not in wb.sheetnames:
        #Create worksheet only if it doesn't exists
        wb.create_sheet('Data', 0)
    else:
        #If it exists clear the data
        wb.worksheets[0].delete_rows(1, wb.worksheets[0].max_row)

    
    #Select Data worksheet
    dataSheet = wb.worksheets[0]

    populate_filter_names(dataSheet)

    #Set counter
    counter = 1

    #Iterate over sheets except the Data sheet
    for sheet in wb:
        if 'Data' not in sheet.title:      
            for row in range(2, sheet.max_row + 1):
                counter += 1
                for col in range(1, sheet.max_column + 1):

                    cell = sheet.cell(row=row, column=col)

                    if sheet.cell(row=1, column=col).value != None:
                        if sheet.cell(row=1, column=col).value.lower() == 'no.':
                            col_num = 2
                            populate_cells(dataSheet, counter, col_num, sheet, cell)
                        elif sheet.cell(row=1, column=col).value.lower() == 'description':
                            col_num = 3
                            populate_cells(dataSheet, counter, col_num, sheet, cell)
                        elif sheet.cell(row=1, column=col).value.lower() == 'unit':
                            col_num = 4
                            populate_cells(dataSheet, counter, col_num, sheet, cell)
                        elif sheet.cell(row=1, column=col).value.lower() == 'boq qty':
                            col_num = 5
                            populate_cells(dataSheet, counter, col_num, sheet, cell)
                        elif sheet.cell(row=1, column=col).value.lower() == 'unit price':
                            col_num = 6
                            populate_cells(dataSheet, counter, col_num, sheet, cell)
                        elif sheet.cell(row=1, column=col).value.lower() == 'amount':
                            col_num = 7
                            populate_cells(dataSheet, counter, col_num, sheet, cell)
                        elif sheet.cell(row=1, column=col).value.lower() == 'prev qty':
                            col_num = 8
                            populate_cells(dataSheet, counter, col_num, sheet, cell)
                        elif sheet.cell(row=1, column=col).value.lower() == 'this qty':
                            col_num = 9
                            populate_cells(dataSheet, counter, col_num, sheet, cell)
                        elif sheet.cell(row=1, column=col).value.lower() == 'total qty':
                            col_num = 10
                            populate_cells(dataSheet, counter, col_num, sheet, cell)
                        elif sheet.cell(row=1, column=col).value.lower() == 'prev amount':
                            col_num = 11
                            populate_cells(dataSheet, counter, col_num, sheet, cell)
                        elif sheet.cell(row=1, column=col).value.lower() == 'this amount':
                            col_num = 12
                            populate_cells(dataSheet, counter, col_num, sheet, cell)
                        elif sheet.cell(row=1, column=col).value.lower() == 'cumulative':
                            col_num = 13
                            populate_cells(dataSheet, counter, col_num, sheet, cell)
                        elif sheet.cell(row=1, column=col).value.lower() == '%':
                            col_num = 14
                            populate_cells(dataSheet, counter, col_num, sheet, cell)
                        elif sheet.cell(row=1, column=col).value.lower() == 'custom1':
                            col_num = 15
                            populate_cells(dataSheet, counter, col_num, sheet, cell)
                        elif sheet.cell(row=1, column=col).value.lower() == 'custom2':
                            col_num = 16
                            populate_cells(dataSheet, counter, col_num, sheet, cell)
                        elif sheet.cell(row=1, column=col).value.lower() == 'custom3':
                            col_num = 17
                            populate_cells(dataSheet, counter, col_num, sheet, cell)
                        elif sheet.cell(row=1, column=col).value.lower() == 'custom4':
                            col_num = 18
                            populate_cells(dataSheet, counter, col_num, sheet, cell)
                        elif sheet.cell(row=1, column=col).value.lower() == 'custom5':
                            col_num = 19
                            populate_cells(dataSheet, counter, col_num, sheet, cell)
                        elif sheet.cell(row=1, column=col).value.lower() == 'custom6':
                            col_num = 20
                            populate_cells(dataSheet, counter, col_num, sheet, cell)
                        elif sheet.cell(row=1, column=col).value.lower() == 'price form':
                            col_num = 21
                            populate_cells(dataSheet, counter, col_num, sheet, cell)                      
                        elif sheet.cell(row=1, column=col).value.lower() == 'contract or vo':
                            col_num = 22
                            populate_cells(dataSheet, counter, col_num, sheet, cell)
                        elif sheet.cell(row=1, column=col).value.lower() == 'item':
                            col_num = 23
                            populate_cells(dataSheet, counter, col_num, sheet, cell)
                        elif sheet.cell(row=1, column=col).value.lower() == 'object':
                            col_num = 24                        
                            populate_cells(dataSheet, counter, col_num, sheet, cell)

                    #Create cells for the sheet names
                    dataSheet.cell(counter, 1).value = sheet.title        

    # Save the wb
    save_wb(wb, filename)  

if __name__ == "__main__":
    main()

    
    
